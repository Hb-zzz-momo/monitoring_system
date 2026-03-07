from __future__ import annotations

import argparse
import math
import re
import sqlite3
import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path
from statistics import mean, pstdev

from openpyxl import load_workbook

BACKEND_DIR = Path(__file__).resolve().parents[1]
REPO_ROOT = BACKEND_DIR.parent
if str(BACKEND_DIR) not in sys.path:
    sys.path.insert(0, str(BACKEND_DIR))

import database as db

DB_PATH = BACKEND_DIR / "monitoring.db"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import exp_data xlsx files into training_samples and sensor_readings")
    parser.add_argument("--data-dir", default=str(REPO_ROOT / "exp_data"), help="Directory containing xlsx files")
    parser.add_argument("--source", default="exp_xlsx", help="training_samples source field")
    parser.add_argument("--append", action="store_true", help="Append instead of replacing existing samples of same source")
    parser.add_argument(
        "--skip-sensor-import",
        action="store_true",
        help="Only generate training samples, skip sensor_readings import",
    )
    return parser.parse_args()


def _safe_float(value: object) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    try:
        return float(text)
    except Exception:
        return None


def _detect_component(name: str) -> str:
    upper_name = name.upper()
    if "IGBT" in upper_name:
        return "IGBT"
    if "MOSFET" in upper_name:
        return "MOSFET"
    return "UNKNOWN"


def _detect_khz(name: str) -> int | None:
    matched = re.search(r"(\d+)\s*k", name, flags=re.IGNORECASE)
    if matched:
        return int(matched.group(1))
    return None


def _detect_noise(name: str) -> bool:
    lower_name = name.lower()
    return "noise" in lower_name or "噪" in lower_name or "song" in lower_name


def _assign_device_id(file_name: str) -> str:
    component = _detect_component(file_name)
    has_noise = _detect_noise(file_name)
    if component == "MOSFET":
        return "2" if has_noise else "1"
    if component == "IGBT":
        return "4" if has_noise else "3"
    return "4"


def _numeric_columns_from_sheet(file_path: Path) -> list[list[float]]:
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []

    # If first row is mostly strings, treat as header and skip it.
    first_row_values = rows[0]
    first_row_numeric_count = sum(1 for v in first_row_values if _safe_float(v) is not None)
    first_row_text_count = sum(1 for v in first_row_values if isinstance(v, str) and v.strip())
    start_index = 1 if first_row_text_count >= first_row_numeric_count else 0

    col_count = max(len(r) for r in rows)
    cols: list[list[float]] = [[] for _ in range(col_count)]

    for row in rows[start_index:]:
        for idx in range(col_count):
            value = row[idx] if idx < len(row) else None
            as_float = _safe_float(value)
            if as_float is not None:
                cols[idx].append(as_float)

    return [c for c in cols if len(c) >= 10]


def _build_sample(file_path: Path) -> tuple[str, str] | None:
    cols = _numeric_columns_from_sheet(file_path)
    if not cols:
        return None

    signal = cols[-1] if len(cols) >= 2 else cols[0]
    if len(signal) < 10:
        return None

    n = len(signal)
    signal_mean = mean(signal)
    signal_std = pstdev(signal) if n > 1 else 0.0
    signal_min = min(signal)
    signal_max = max(signal)
    first_value = signal[0]
    last_value = signal[-1]
    slope = (last_value - first_value) / max(1, n - 1)

    component = _detect_component(file_path.name)
    khz = _detect_khz(file_path.stem)
    has_noise = _detect_noise(file_path.stem)

    freq_desc = f"{khz}kHz" if khz is not None else "未知频段"
    noise_desc = "含噪声" if has_noise else "常规波形"

    input_text = (
        f"实验文件: {file_path.name}。"
        f"器件类型: {component}。"
        f"频段: {freq_desc}。"
        f"数据属性: {noise_desc}。"
        f"样本点数: {n}。"
        f"电压统计: 均值={signal_mean:.6f}, 标准差={signal_std:.6f}, "
        f"最小值={signal_min:.6f}, 最大值={signal_max:.6f}, "
        f"首值={first_value:.6f}, 末值={last_value:.6f}, 趋势斜率={slope:.8f}。"
        "请给出该样本的状态判断和运维建议。"
    )

    trend = "上升" if slope > 0.0001 else "下降" if slope < -0.0001 else "平稳"
    fluctuation = "较大" if signal_std > 0.15 else "中等" if signal_std > 0.05 else "较小"

    expected_output = (
        f"判断: 该样本属于{component}器件在{freq_desc}下的{noise_desc}信号，"
        f"整体波动{fluctuation}，趋势{trend}。"
        "建议: 1) 持续跟踪该频段同类样本均值与方差；"
        "2) 若标准差持续扩大或均值漂移加剧，安排器件热-电联合复测；"
        "3) 对含噪声样本先做滤波和复采样，再进行故障判定。"
    )

    return input_text, expected_output


def _build_sensor_payloads(file_path: Path) -> list[dict[str, float | int | bool | str]]:
    cols = _numeric_columns_from_sheet(file_path)
    if not cols:
        return []

    signal = cols[-1] if len(cols) >= 2 else cols[0]
    if len(signal) < 10:
        return []

    signal_min = min(signal)
    signal_max = max(signal)
    signal_range = max(1e-6, signal_max - signal_min)
    component = _detect_component(file_path.name)
    has_noise = _detect_noise(file_path.name)
    device_id = _assign_device_id(file_path.name)

    start_time = datetime.now(timezone.utc) - timedelta(seconds=len(signal))
    payloads: list[dict[str, float | int | bool | str]] = []

    for idx, value in enumerate(signal):
        normalized = max(0.0, min(1.0, (value - signal_min) / signal_range))

        voltage = 205.0 + normalized * 30.0
        current = 8.0 + normalized * 14.0
        power = 1.8 + normalized * 3.2 + (0.35 if component == "IGBT" else 0.0)
        temperature = 34.0 + normalized * 44.0 + (2.2 if has_noise else 0.0)
        energy = 120.0 + idx * 0.035 + normalized * 3.5
        delay = 8 + int(abs((normalized - 0.5) * 20))

        payloads.append(
            {
                "deviceId": device_id,
                "timestamp": (start_time + timedelta(seconds=idx)).isoformat(timespec="seconds"),
                "temperature": float(f"{temperature:.3f}"),
                "voltage": float(f"{voltage:.3f}"),
                "current": float(f"{current:.3f}"),
                "power": float(f"{power:.3f}"),
                "energy": float(f"{energy:.3f}"),
                "delay": delay,
                "isConnected": True,
            }
        )

    return payloads


def _replace_source_samples(source: str) -> None:
    conn = sqlite3.connect(DB_PATH)
    try:
        conn.execute("DELETE FROM training_samples WHERE source = ?", (source,))
        conn.commit()
    finally:
        conn.close()


def _reset_sensor_readings() -> None:
    conn = sqlite3.connect(DB_PATH)
    try:
        conn.execute("DELETE FROM sensor_readings")
        conn.commit()
    finally:
        conn.close()


def main() -> None:
    args = parse_args()
    data_dir = Path(args.data_dir)
    if not data_dir.exists():
        raise FileNotFoundError(f"Data dir not found: {data_dir}")

    files = sorted(data_dir.glob("*.xlsx"))
    if not files:
        raise FileNotFoundError(f"No xlsx files found in: {data_dir}")

    if not args.append:
        _replace_source_samples(args.source)
        if not args.skip_sensor_import:
            _reset_sensor_readings()

    imported = 0
    sensor_rows = 0
    skipped: list[str] = []

    for file_path in files:
        sample = _build_sample(file_path)
        if sample is None:
            skipped.append(file_path.name)
            continue
        input_text, expected_output = sample
        db.add_training_sample(args.source, input_text, expected_output)
        imported += 1

        if not args.skip_sensor_import:
            payloads = _build_sensor_payloads(file_path)
            for payload in payloads:
                db.ingest_sensor_reading(payload, add_training_sample=False)
                sensor_rows += 1

    print(f"Imported {imported} samples from {len(files)} files with source={args.source}")
    if not args.skip_sensor_import:
        print(f"Imported {sensor_rows} sensor points into sensor_readings")
    if skipped:
        print("Skipped files:")
        for name in skipped:
            print(f"- {name}")


if __name__ == "__main__":
    main()
